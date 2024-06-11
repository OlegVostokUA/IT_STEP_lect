import os, sys
import openpyxl
import locale
from datetime import datetime


def scrap_file_2():
    # src_file = self.get_src_file_path()
    src_file = '94 ПРИКЗ Залишки продуктів (по підрозділах) 07.05.2024.xlsx'
    # base_file = self.get_file_path()
    base_file = '94 ПРИКЗ Орган охорони ДК 08.05.2024 – копія.xlsx'
    # read data from source file
    read_data_src_file = openpyxl.load_workbook(src_file)
    sheet_1_src_file = read_data_src_file.sheetnames[0]
    sheet_src_file = read_data_src_file[sheet_1_src_file]
    # columns = [32, 33]
    dict_source_1 = {}
    dict_source_2 = {}

    for row in range(7, 119):
        key = sheet_src_file.cell(column=1, row=row).value
        val_1 = sheet_src_file.cell(column=32, row=row).value
        val_2 = sheet_src_file.cell(column=33, row=row).value
        dict_source_1[key] = val_1
        dict_source_2[key] = val_2

    read_data_base_file = openpyxl.load_workbook(base_file)
    sheet_1_base_file = read_data_base_file.sheetnames[0]
    sheet_base_file = read_data_base_file[sheet_1_base_file]
    ### 1
    for item in dict_source_1.items():
        name, value = item[0], item[1]

        prev_key = None
        for i in range(6, 117):
            key = sheet_base_file.cell(column=1, row=i).value

            if key == None:
                key = prev_key
            else:
                prev_key = key

            if key in ['Капуста', 'Морква', 'Буряк', 'Цибуля', 'Часник', 'Огірки, помідори, баклажани']:

                key = key + ' ' + sheet_base_file.cell(column=2, row=i).value

            if key.lower() == name.lower():
                row = sheet_base_file.cell(column=1, row=i).row
                sheet_base_file.cell(column=3, row=row).value = value
                break
    ### 2
    for item in dict_source_2.items():
        name, value = item[0], item[1]

        prev_key = None
        for i in range(6, 117):
            key = sheet_base_file.cell(column=1, row=i).value

            if key == None:
                key = prev_key
            else:
                prev_key = key

            if key in ['Капуста', 'Морква', 'Буряк', 'Цибуля', 'Часник', 'Огірки, помідори, баклажани']:

                key = key + ' ' + sheet_base_file.cell(column=2, row=i).value

            if key.lower() == name.lower():
                row = sheet_base_file.cell(column=1, row=i).row
                sheet_base_file.cell(column=5, row=row).value = value
                break

    new_filename = base_file.split('.')
    new_filename.insert(1, date)
    new_filename[2] = '.xlsx'
    new_filename = ' '.join(new_filename)
    read_data_base_file.save(new_filename)

scrap_file_2()